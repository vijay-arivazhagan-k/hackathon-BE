"""
Request API endpoints for FastAPI
"""
from fastapi import APIRouter, HTTPException, status, Query
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional
from schemas.requests import (
    RequestCreate, RequestResponse, RequestStatusUpdate,
    PaginatedRequests, InsightsResponse, RequestHistoryResponse
)
from services.request_service import RequestService
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


router = APIRouter(prefix="/api/requests", tags=["requests"])
service = RequestService()


def serialize_response(model_or_list, exclude_aliases=True):
    """Helper to serialize Pydantic models without aliases (using field names)"""
    if isinstance(model_or_list, list):
        return [m.model_dump(by_alias=False, exclude_none=False) if hasattr(m, 'model_dump') else m 
                for m in model_or_list]
    elif hasattr(model_or_list, 'model_dump'):
        return model_or_list.model_dump(by_alias=False, exclude_none=False)
    return model_or_list


@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_request(request: RequestCreate):
    """
    Create a new request
    
    Args:
        request: Request creation data
        
    Returns:
        RequestResponse: Created request
    """
    req = service.create_request(
        request.user_id, request.total_amount, request.invoice_date,
        request.invoice_number, request.category_name, request.comments,
        request.approvaltype if request.approvaltype else 'Auto'
    )
    
    response = RequestResponse.model_validate(req)
    return JSONResponse(content=serialize_response(response), status_code=201)


@router.get("/", response_model=PaginatedRequests)
async def list_requests(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    status: Optional[str] = Query(default=None, description="Filter by status: Pending, Approved, Rejected, All"),
    start: Optional[str] = Query(default=None, description="Start creation date filter (YYYY-MM-DD)"),
    end: Optional[str] = Query(default=None, description="End creation date filter (YYYY-MM-DD)"),
    category_id: Optional[str] = Query(default=None, description="Category name filter")
):
    """List all requests with pagination and filters

    Date range filters now apply to the request creation timestamp (`CREATED_ON`) rather than invoice date.

    Args:
        page: Page number (default: 1)
        page_size: Items per page (default: 20, max: 100)
        status: Status filter (Pending, Approved, Rejected, All)
        start: Start creation date filter (YYYY-MM-DD)
        end: End creation date filter (YYYY-MM-DD)
        category_id: Category name filter

    Returns:
        PaginatedRequests: Paginated requests
    """
    print(f"🔍 API received filters: page={page}, page_size={page_size}, status={status}, start={start}, end={end}, category_id={category_id}")
    requests, total = service.list_requests(page, page_size, status, start, end, category_id)
    
    print(f"✅ API got {len(requests)} requests from service, total={total}")
    
    # Convert to response format - returns camelCase field names
    items = [RequestResponse.model_validate(req) for req in requests]
    
    print(f"✅ API converted to {len(items)} response items")
    
    response_data = {
        "items": serialize_response(items),
        "page": page,
        "page_size": page_size,
        "total": total
    }
    
    print(f"✅ API returning: {len(response_data['items'])} items, page={page}, total={total}")
    
    return JSONResponse(content=response_data)


@router.get("/export", response_class=StreamingResponse)
async def export_requests(
    start: Optional[str] = Query(default=None, description="Start creation date (YYYY-MM-DD)"),
    end: Optional[str] = Query(default=None, description="End creation date (YYYY-MM-DD)"),
    category_id: Optional[str] = Query(default=None, description="Category name"),
    status: Optional[str] = Query(default=None, description="Filter by status: Pending, Approved, Rejected, All")
):
    """Export requests as Excel file

    Date range filters are applied to `CREATED_ON` (request creation timestamp).

    Args:
        start: Start creation date filter (YYYY-MM-DD)
        end: End creation date filter (YYYY-MM-DD)
        category_id: Category filter
        status: Status filter

    Returns:
        StreamingResponse: Excel file
    """
    try:
        # Get filtered requests
        requests = service.get_filtered_requests_for_export(start, end, category_id, status)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requests"
        
        # Define headers
        headers = [
            "ID", "User ID", "Total Amount (₹)", "Approved Amount (₹)", "Invoice Date", "Invoice Number",
            "Category", "Status", "Comments", "Approval Type", "Created On", "Created By"
        ]
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_num, req in enumerate(requests, 2):
            ws.cell(row=row_num, column=1).value = req.ID
            ws.cell(row=row_num, column=2).value = req.USER_ID
            ws.cell(row=row_num, column=3).value = float(req.TOTAL_AMOUNT) if req.TOTAL_AMOUNT else 0
            ws.cell(row=row_num, column=4).value = float(req.APPROVED_AMOUNT) if req.APPROVED_AMOUNT else 0
            ws.cell(row=row_num, column=5).value = req.INVOICE_DATE
            ws.cell(row=row_num, column=6).value = req.INVOICE_NUMBER
            ws.cell(row=row_num, column=7).value = req.CATEGORY_NAME
            ws.cell(row=row_num, column=8).value = req.CURRENT_STATUS
            ws.cell(row=row_num, column=9).value = req.COMMENTS
            ws.cell(row=row_num, column=10).value = req.APPROVALTYPE
            ws.cell(row=row_num, column=11).value = req.CREATED_ON
            ws.cell(row=row_num, column=12).value = req.CREATED_BY
            
            # Format currency columns
            ws.cell(row=row_num, column=3).number_format = '"₹"#,##0.00'
            ws.cell(row=row_num, column=4).number_format = '"₹"#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 15
        
        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"requests_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting requests: {str(e)}"
        )


@router.get("/{request_id}", response_model=RequestResponse)
async def get_request(request_id: int):
    """
    Get request by ID
    
    Args:
        request_id: Request ID
        
    Returns:
        RequestResponse: Request details
        
    Raises:
        HTTPException: 404 if request not found
    """
    req = service.get_request(request_id)
    if not req:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Request with ID {request_id} not found"
        )
    
    response = RequestResponse.model_validate(req)
    return JSONResponse(content=serialize_response(response))


@router.patch("/{request_id}/status", response_model=RequestResponse)
async def update_request_status(request_id: int, update: RequestStatusUpdate):
    """
    Update request status
    
    Args:
        request_id: Request ID
        update: Status update data
        
    Returns:
        RequestResponse: Updated request
        
    Raises:
        HTTPException: 404 if request not found
    """
    req = service.update_request_status(
        request_id, update.status, update.comments, 
        update.updated_by if update.updated_by else 'Admin',
        update.approved_amount
    )
    
    if not req:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Request with ID {request_id} not found"
        )
    
    response = RequestResponse.model_validate(req)
    return JSONResponse(content=serialize_response(response))


@router.get("/{request_id}/history", response_model=list[RequestHistoryResponse])
async def get_request_history(request_id: int):
    """
    Get request history
    
    Args:
        request_id: Request ID
        
    Returns:
        List[RequestHistoryResponse]: Request history
    """
    # First check if request exists
    req = service.get_request(request_id)
    if not req:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Request with ID {request_id} not found"
        )
    
    history = service.get_request_history(request_id)
    
    responses = [RequestHistoryResponse.model_validate(h) for h in history]
    return JSONResponse(content=serialize_response(responses))


@router.get("/insights/summary", response_model=InsightsResponse)
async def get_insights(
    start: Optional[str] = Query(default=None, description="Start creation date filter (YYYY-MM-DD)"),
    end: Optional[str] = Query(default=None, description="End creation date filter (YYYY-MM-DD)"),
    duration: Optional[str] = Query(default=None, description="Duration filter (deprecated, use start/end)")
):
    """Get request insights/statistics

    Date range filters are applied to `CREATED_ON`.

    Args:
        start: Start creation date filter (YYYY-MM-DD)
        end: End creation date filter (YYYY-MM-DD)
        duration: Duration filter (deprecated)

    Returns:
        InsightsResponse: Insights data
    """
    insights = service.get_insights(start_date=start, end_date=end, duration=duration)
    return JSONResponse(content={
        "total": insights['total'],
        "approved": insights['approved'],
        "rejected": insights['rejected'],
        "pending": insights['pending'],
        "status_breakdown": insights['status_breakdown']
    })